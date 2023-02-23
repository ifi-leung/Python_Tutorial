# import openpyxl

# workbook = openpyxl.Workbook()
# sheet = workbook.active

# sheet['A1'] = 'hello pytho excel !'
# workbook.save(filename="./test.xlsx")
################
# import openpyxl

# workbook = openpyxl.load_workbook("./test.xlsx")
# sheet = workbook.active

# a2 = sheet.cell(row = 2, column = 1)
# a3 = sheet['A3']

# a2.value = "修改属性1"
# a3.value = "修改属性2"
# sheet['A4'] = "直接赋值"

# workbook.save(filename="./test.xlsx")
################
# import openpyxl 
  
# wb = openpyxl.load_workbook("test.xlsx")
# sheet = wb.active

# data = (
#     ('a', 'b', 'c'),
#     ('d', 'e', 'f')
# )  
# for row in data:
#     sheet.append(row)

# wb.save('test.xlsx')
################
# import openpyxl

# workbook = openpyxl.load_workbook("test.xlsx") 
# sheet = workbook.active 
  
# row = sheet.max_row
# column = sheet.max_column
# print(f"Total Row num: {row}")
# print(f"Total Column num: {column}")

# print("\nValue of first column:")
# for i in range(1, row + 1):
#     print(sheet.cell(row = i, column = 1).value)

# print("\nValue of second row")
# for i in range(1, column + 1):
#     print(sheet.cell(row = 2, column = i).value, end = " ")
################
# import openpyxl

# workbook = openpyxl.load_workbook("test.xlsx") 
# sheet = workbook.active 

# cell_obj = sheet['A5:C6']
# for cell1, cell2, cell3 in cell_obj:
#     print(cell1.value,
#             cell2.value,
#             cell3.value)
################
# import openpyxl

# workbook = openpyxl.Workbook()
# sheet = workbook.active

# sheet['A1'] = 100
# sheet['A2'] = 200
# sheet['A3'] = 300
# sheet['A4'] = 400
# sheet['A5'] = 500

# sheet['A7'] = '= SUM(A1:A5)'

# workbook.save("sum.xlsx")
################
# import openpyxl
  
# workbook = openpyxl.load_workbook("test.xlsx")
# sheet = workbook.active

# sheet.row_dimensions[6].height = 70
# sheet.column_dimensions['C'].width = 20
  
# workbook.save('test.xlsx')
################
# import openpyxl
  
# workbook = openpyxl.load_workbook("test.xlsx")
# sheet = workbook.active

# sheet.merge_cells("B2:C4")
  
# workbook.save('test.xlsx')
################
# import openpyxl
  
# workbook = openpyxl.load_workbook("test.xlsx")
# sheet = workbook.active
# sheet.unmerge_cells("B2:C4")
  
# workbook.save('test.xlsx')
################
# import openpyxl
# from openpyxl.styles import Font
  
# workbook = openpyxl.load_workbook("test.xlsx")
# sheet = workbook.active

# sheet.cell(row = 1,
#             column = 1).font = Font(size = 24,
#                                     italic = True,
#                                     bold = True,
#                                     name = 'Times New Roman')
  
# workbook.save('test.xlsx')
################
# import openpyxl
# from openpyxl.chart import BarChart
# from openpyxl.chart import Reference

# workbook = openpyxl.Workbook()
# sheet = workbook.active

# # 插入数据到表格 sheet
# for i in range(10):
# 	sheet.append([i])
# # 选中表格的数据域
# values = Reference(sheet,
#                     min_col = 1,
#                     min_row = 1,
#                     max_col = 1,
#                     max_row = 10)

# # 创建柱状图表
# chart = BarChart()
# # 关联数据域到图表
# chart.add_data(values)
# # 设置图表表头
# chart.title = " 实例图标 "
# chart.x_axis.title = " X 轴 "
# chart.y_axis.title = " Y 轴 "
# # 生成图表，并贴到单元格 C2
# sheet.add_chart(chart, "C2")

# workbook.save("sample.xlsx")
################
import openpyxl
from openpyxl.drawing.image import Image

workbook = openpyxl.Workbook()
sheet = workbook.active

img = Image("beauty.jpg")
sheet.add_image(img, 'B2')

workbook.save("img.xlsx")

